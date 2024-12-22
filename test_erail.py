import datetime
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from HomePage import HomePage


def test_erail():
    # Initialize the Chrome WebDriver
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

    try:
        # Initialize the ERAILHomePage object
        homepage = HomePage(driver)

        # Step 1: Open URL
        homepage.open()

        # Step 2: Click on "From" field
        homepage.click_from_station()

        # Step 3: Clear the "From" field
        homepage.clear_from_station()

        # Step 4: Insert data "DEL" to open the dropdown
        homepage.set_from_station("DEL")

        # Step 5: Select the station at the 4th position and print it
        station_name = homepage.select_4th_station()
        print(f"Selected station: {station_name}")

        # Step 6: Create an Excel file and write expected station names
        workbook = Workbook()
        sheet = workbook.active
        expected_stations = ["New Delhi", "Delhi Cantt", "Delhi Sarai Rohilla", "Delhi"]
        sheet.append(["Expected Stations"])
        for station in expected_stations:
            sheet.append([station])
        workbook.save("expected_stations.xlsx")

        # Step 7: Get the dropdown list data and write it into the Excel file, then compare
        dropdown_data = homepage.get_dropdown_data() # Load the expected stations workbook
        expected_workbook = load_workbook("expected_stations.xlsx")
        expected_sheet = expected_workbook.active # Write the dropdown data into a new sheet
        dropdown_sheet = expected_workbook.create_sheet(title="Dropdown Data")
        dropdown_sheet.append(["Dropdown Stations"])
        for station in dropdown_data: dropdown_sheet.append([station]) # Compare the dropdown data with expected data
        expected_set = set(expected_stations)
        dropdown_set = set(dropdown_data)
        print("Expected stations:", expected_set)
        print("Dropdown stations:", dropdown_set)
        if expected_set == dropdown_set: print("Dropdown data matches the expected stations.")
        else: print("Dropdown data does not match the expected stations.") # Save the workbook with dropdown data
        expected_workbook.save("station_comparison.xlsx")

        # Step 8: Select 30 days from the current date in "Sort on Date"
        current_date = datetime.datetime.now()
        future_date = (current_date + datetime.timedelta(days=30)).strftime("%d-%m-%Y")
        homepage.set_date(future_date)

        print("Test completed successfully.")

    finally:
        driver.quit()


# Run the test case
test_erail()
